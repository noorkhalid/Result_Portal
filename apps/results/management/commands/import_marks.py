from django.core.management.base import BaseCommand
from django.db import transaction
import openpyxl

from academics.models import Program, Session, Course
from students.models import Student, Enrollment
from results.models import ResultBatch, CourseResult
from results.services import recompute_batch


def _norm(s: str) -> str:
    return "".join(ch.lower() for ch in str(s).strip() if ch.isalnum())


def _to_decimal_or_zero(v):
    if v is None or str(v).strip() == "":
        return 0
    return float(v)


class Command(BaseCommand):
    help = "Import marks Excel and create CourseResult + compute GPA/CGPA."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")
        parser.add_argument("--recompute", action="store_true", help="Recompute batch GPA/CGPA after import")

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options["file"]

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        header_raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header = [_norm(h) for h in header_raw]

        def col(*names):
            for n in names:
                n2 = _norm(n)
                for i, h in enumerate(header):
                    if h == n2:
                        return i
            return None

        reg_i = col("registration_no", "registrationno")
        prog_i = col("program")
        sess_i = col("session")
        sem_i = col("semester")
        code_i = col("course_code", "coursecode", "code")
        title_i = col("course_title", "coursetitle", "title")

        sesm_i = col("sessional_marks", "sessional")
        mid_i = col("midterm_marks", "midterm", "mid")
        ter_i = col("terminal_marks", "terminal", "final")
        max_i = col("maxmarks", "max_marks", "max")

        exam_i = col("examtype", "result_type", "type")

        required = {
            "registration_no": reg_i,
            "program": prog_i,
            "session": sess_i,
            "semester": sem_i,
            "terminal_marks": ter_i,
            "maxmarks": max_i,
        }
        missing = [k for k, v in required.items() if v is None]
        if missing:
            raise SystemExit(f"Missing required columns: {missing}\nHeaders found: {header_raw}")

        created = 0
        updated = 0
        skipped = 0
        errors = 0

        batches = {}  # cache ResultBatch by (program_id, session_id, semester, type)

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                registration_no = row[reg_i]
                program_name = row[prog_i]
                session_year = row[sess_i]
                semester_number = row[sem_i]
                terminal_marks = row[ter_i]
                max_marks = row[max_i]

                # optional
                course_code = row[code_i] if code_i is not None else None
                course_title = row[title_i] if title_i is not None else None
                sessional_marks = row[sesm_i] if sesm_i is not None else None
                midterm_marks = row[mid_i] if mid_i is not None else None
                examtype = row[exam_i] if exam_i is not None else "Regular"

                if not registration_no or not program_name or not session_year or not semester_number:
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"Row {row_num}: missing program/session/semester/registration_no"))
                    continue

                if terminal_marks is None or str(terminal_marks).strip() == "":
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"Row {row_num}: terminal_marks is required"))
                    continue

                if max_marks is None or str(max_marks).strip() == "":
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"Row {row_num}: maxmarks is required"))
                    continue

                registration_no = str(registration_no).strip()
                program_name = str(program_name).strip()
                session_year = int(session_year)
                semester_number = int(semester_number)

                examtype = str(examtype).strip().lower()
                if examtype in ("regular",):
                    result_type = "regular"
                elif examtype in ("repeat", "reappear"):
                    result_type = "repeat"
                elif examtype in ("improved", "improvement"):
                    result_type = "improved"
                else:
                    # fallback
                    result_type = "regular"

                program = Program.objects.filter(name=program_name).first()
                if not program:
                    # fuzzy match
                    program = Program.objects.filter(name__icontains=program_name).first()
                if not program:
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"Row {row_num}: Program not found: {program_name}"))
                    continue

                session = Session.objects.filter(start_year=session_year).first()
                if not session:
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"Row {row_num}: Session not found: {session_year}"))
                    continue

                student = Student.objects.filter(registration_no=registration_no).first()
                if not student:
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"Row {row_num}: Student not found: {registration_no}"))
                    continue

                enrollment = Enrollment.objects.filter(student=student, program=program, session=session).first()
                if not enrollment:
                    errors += 1
                    self.stdout.write(self.style.ERROR(
                        f"Row {row_num}: Enrollment not found for reg={registration_no} program={program.name} session={session.start_year}"
                    ))
                    continue

                # Course match: by code preferred, else title
                course = None
                if course_code is not None and str(course_code).strip() != "":
                    course = Course.objects.filter(code=str(course_code).strip()).first()

                if not course and course_title:
                    course = Course.objects.filter(title=str(course_title).strip()).first()

                if not course:
                    errors += 1
                    self.stdout.write(self.style.ERROR(
                        f"Row {row_num}: Course not found (code={course_code}, title={course_title})"
                    ))
                    continue

                key = (program.id, session.id, semester_number, result_type)
                if key not in batches:
                    batch, _ = ResultBatch.objects.get_or_create(
                        program=program,
                        session=session,
                        semester_number=semester_number,
                        result_type=result_type,
                    )
                    batches[key] = batch
                batch = batches[key]

                s_marks = _to_decimal_or_zero(sessional_marks)
                m_marks = _to_decimal_or_zero(midterm_marks)
                t_marks = _to_decimal_or_zero(terminal_marks)

                total_marks = s_marks + m_marks + t_marks

                cr = CourseResult.objects.filter(batch=batch, enrollment=enrollment, course=course).first()
                if not cr:
                    CourseResult.objects.create(
                        batch=batch,
                        enrollment=enrollment,
                        course=course,
                        marks_obtained=total_marks,
                        max_marks=float(max_marks),
                    )
                    created += 1
                else:
                    cr.marks_obtained = total_marks
                    cr.max_marks = float(max_marks)
                    cr.save(update_fields=["marks_obtained", "max_marks"])
                    updated += 1

            except Exception as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {row_num}: ERROR {e}"))

        self.stdout.write(self.style.SUCCESS(
            f"\nImported. Created={created}, Updated={updated}, Errors={errors}"
        ))

        if options["recompute"]:
            for batch in batches.values():
                recompute_batch(batch)
            self.stdout.write(self.style.SUCCESS("Recompute done (GPA/CGPA updated)."))
        else:
            self.stdout.write(self.style.WARNING("Recompute skipped. Run with --recompute to calculate GPA/CGPA."))
