from django.core.management.base import BaseCommand
from results.models import GradeScale
import openpyxl


def _norm(s: str) -> str:
    return "".join(ch.lower() for ch in str(s).strip() if ch.isalnum())


class Command(BaseCommand):
    help = "Import GradeScale from Excel (minpercent, maxpercent, lettergrade, gradepoint, remarks)."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")
        parser.add_argument("--clear", action="store_true", help="Delete existing grade scale rows first")

    def handle(self, *args, **options):
        file_path = options["file"]

        if options["clear"]:
            GradeScale.objects.all().delete()
            self.stdout.write(self.style.WARNING("Cleared existing GradeScale rows."))

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        header_raw = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        header = [_norm(h) for h in header_raw]

        def find_col(cands):
            for c in cands:
                cn = _norm(c)
                for i, h in enumerate(header):
                    if h == cn:
                        return i
            return None

        min_i = find_col(["minpercent", "min_percentage", "min", "from"])
        max_i = find_col(["maxpercent", "max_percentage", "max", "to"])
        let_i = find_col(["lettergrade", "letter_grade", "grade", "letter"])
        gp_i = find_col(["gradepoint", "grade_point", "gp"])
        rem_i = find_col(["remarks", "remark"])

        missing = []
        if min_i is None: missing.append("minpercent")
        if max_i is None: missing.append("maxpercent")
        if let_i is None: missing.append("lettergrade")
        if gp_i is None: missing.append("gradepoint")
        if rem_i is None: missing.append("remarks")

        if missing:
            raise SystemExit(f"Missing columns: {missing}\nHeaders found: {header_raw}")

        created = 0
        skipped = 0
        errors = 0

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            min_p = row[min_i]
            max_p = row[max_i]
            letter = row[let_i]
            gp = row[gp_i]
            remarks = row[rem_i]

            if min_p is None or max_p is None or letter is None or gp is None:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {row_num}: missing values (skipped)"))
                continue

            letter = str(letter).strip()
            remarks = str(remarks).strip() if remarks is not None else "Pass"

            is_fail = "fail" in remarks.lower() or float(gp) == 0

            obj, is_created = GradeScale.objects.get_or_create(
                min_percentage=float(min_p),
                max_percentage=float(max_p),
                defaults={
                    "letter_grade": letter,
                    "grade_point": float(gp),
                    "remarks": remarks,
                    "is_fail": is_fail,
                }
            )

            if is_created:
                created += 1
            else:
                # update existing row
                obj.letter_grade = letter
                obj.grade_point = float(gp)
                obj.remarks = remarks
                obj.is_fail = is_fail
                obj.save(update_fields=["letter_grade", "grade_point", "remarks", "is_fail"])
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"Done. Created={created}, Updated={skipped}, Errors={errors}"
        ))
