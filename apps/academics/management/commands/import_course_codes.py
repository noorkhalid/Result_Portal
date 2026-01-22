from django.core.management.base import BaseCommand
from academics.models import Course
import openpyxl


def _norm(s: str) -> str:
    return " ".join(str(s).strip().split()).lower()


class Command(BaseCommand):
    help = "Import/Update Course.code by matching Course.title from Excel (title, coursecode)."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")
        parser.add_argument("--overwrite", action="store_true", help="Overwrite existing codes if different")

    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(options["file"])
        ws = wb.active

        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

        # Accept your exact headers
        try:
            title_i = headers.index("title")
        except ValueError:
            raise SystemExit(f"Missing column: title. Found headers: {headers}")

        # coursecode column can be called coursecode / code
        code_i = None
        for name in ("coursecode", "code"):
            if name in headers:
                code_i = headers.index(name)
                break
        if code_i is None:
            raise SystemExit(f"Missing column: coursecode (or code). Found headers: {headers}")

        updated = 0
        skipped = 0
        errors = 0

        for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            title = row[title_i]
            code = row[code_i]

            if not title or not code:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {r}: missing title/code"))
                continue

            title_clean = " ".join(str(title).strip().split())
            code_clean = str(code).strip()

            course = Course.objects.filter(title=title_clean).first()
            if not course:
                # try normalized search to handle extra spaces/case
                course = next(
                    (c for c in Course.objects.all() if _norm(c.title) == _norm(title_clean)),
                    None
                )

            if not course:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {r}: Course not found by title: {title_clean}"))
                continue

            if course.code and not options["overwrite"] and course.code != code_clean:
                skipped += 1
                self.stdout.write(self.style.WARNING(
                    f"Skipped (has code): {course.title} | existing={course.code} | new={code_clean}"
                ))
                continue

            if course.code != code_clean:
                course.code = code_clean
                course.save(update_fields=["code"])
                updated += 1
                self.stdout.write(self.style.SUCCESS(f"Updated: {course.title} -> {code_clean}"))
            else:
                skipped += 1

        self.stdout.write(self.style.SUCCESS(f"\nDone. Updated={updated}, Skipped={skipped}, Errors={errors}"))
