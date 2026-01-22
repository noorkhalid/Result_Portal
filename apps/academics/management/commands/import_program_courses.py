from django.core.management.base import BaseCommand
from academics.models import Program, Course, ProgramCourse
import openpyxl


def norm(s):
    return str(s).strip().lower()


class Command(BaseCommand):
    help = "Import ProgramCourse mapping from Excel (program, semester, title)"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        file_path = options["file"]
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        header = [norm(c.value) for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        p_idx = header.index("program")
        s_idx = header.index("semester")
        t_idx = header.index("title")

        created = 0
        skipped = 0
        errors = 0

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            program_name = row[p_idx]
            sem_no = row[s_idx]
            title = row[t_idx]

            if not program_name or not sem_no or not title:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {row_num}: Missing data (skipped)"))
                continue

            try:
                program = Program.objects.get(name=str(program_name).strip())
            except Program.DoesNotExist:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {row_num}: Program not found: {program_name}"))
                continue

            try:
                course = Course.objects.get(title=str(title).strip())
            except Course.DoesNotExist:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {row_num}: Course not found: {title}"))
                continue

            obj, is_created = ProgramCourse.objects.get_or_create(
                program=program,
                semester_number=int(sem_no),
                course=course,
            )

            if is_created:
                created += 1
                self.stdout.write(self.style.SUCCESS(f"Mapped: {program} | Sem {sem_no} | {course.title}"))
            else:
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nDone. Created={created}, Skipped={skipped}, Errors={errors}"
        ))
