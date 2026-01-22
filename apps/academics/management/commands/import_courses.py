from django.core.management.base import BaseCommand
from academics.models import Course
import openpyxl


def _normalize(s: str) -> str:
    return "".join(ch.lower() for ch in str(s).strip() if ch.isalnum())


class Command(BaseCommand):
    help = "Import courses from Excel. Requires title + credit hours."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        file_path = options["file"]

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Read header row and detect columns by name
        header = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_norm = [_normalize(h) for h in header]

        # Find title column
        title_candidates = {"title", "coursetitle", "subject", "name"}
        ch_candidates = {"credithours", "credithours", "credithour", "credithr", "credit_hours", "ch"}

        try:
            title_idx = next(i for i, h in enumerate(header_norm) if h in title_candidates)
        except StopIteration:
            raise SystemExit(f"Could not find a title column in header: {header}")

        try:
            ch_idx = next(i for i, h in enumerate(header_norm) if h in ch_candidates)
        except StopIteration:
            raise SystemExit(f"Could not find a credit hours column in header: {header}")

        created = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            title = row[title_idx] if title_idx < len(row) else None
            credit_hours = row[ch_idx] if ch_idx < len(row) else None

            if not title or not str(title).strip():
                skipped += 1
                continue

            title = str(title).strip()

            # Validate credit hours
            if credit_hours is None or str(credit_hours).strip() == "":
                errors += 1
                self.stdout.write(self.style.ERROR(
                    f"Row {row_num}: Missing credit hours for '{title}' (skipped)"
                ))
                continue

            # Convert credit hours safely
            try:
                credit_hours = float(credit_hours)
            except Exception:
                errors += 1
                self.stdout.write(self.style.ERROR(
                    f"Row {row_num}: Invalid credit hours '{credit_hours}' for '{title}' (skipped)"
                ))
                continue

            # Create or update
            obj, is_created = Course.objects.get_or_create(
                title=title,
                defaults={"credit_hours": credit_hours},
            )

            if is_created:
                created += 1
                self.stdout.write(self.style.SUCCESS(f"Created: {title} ({credit_hours} CH)"))
            else:
                # If existing, update credit hours if different
                if float(obj.credit_hours) != float(credit_hours):
                    obj.credit_hours = credit_hours
                    obj.save(update_fields=["credit_hours"])
                    updated += 1
                    self.stdout.write(self.style.WARNING(f"Updated CH: {title} -> {credit_hours}"))
                else:
                    skipped += 1
                    self.stdout.write(self.style.WARNING(f"Skipped (exists): {title}"))

        self.stdout.write(self.style.SUCCESS(
            f"\nDone. Created={created}, Updated={updated}, Skipped={skipped}, Errors={errors}"
        ))
