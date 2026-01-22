from __future__ import annotations

import os
from datetime import datetime

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.db.models import Count
from django.shortcuts import redirect, render

from academics.models import Course, Program, Session
from results.models import CourseResult, ResultBatch
from results.services import recompute_batch
from students.models import Enrollment, Student

from .decorators import group_required


def home(request):
    """Public home.

    If logged in, send user to their dashboard. Otherwise, send to login.
    """
    if request.user.is_authenticated:
        return redirect("dashboard")
    return redirect("login")


@login_required
def dashboard(request):
    """Route the logged-in user to the correct group dashboard."""
    user = request.user

    # System Admin: either superuser or in "System Admin" group
    if user.is_superuser or user.groups.filter(name="System Admin").exists():
        return redirect("dash_system_admin")

    # Order matters if a user belongs to multiple groups
    if user.groups.filter(name="Controller").exists():
        return redirect("dash_controller")
    if user.groups.filter(name="Data Entry").exists():
        return redirect("dash_data_entry")
    if user.groups.filter(name="Document Generator").exists():
        return redirect("dash_document_generator")
    if user.groups.filter(name="Result Checker").exists():
        return redirect("dash_result_checker")

    # Fallback: no group assigned
    return render(request, "dashboards/no_group.html")


@group_required("Controller")
def controller_dashboard(request):
    return render(request, "dashboards/controller.html")


@group_required("Data Entry")
def data_entry_dashboard(request):
    """Data Entry dashboard with quick stats and links."""
    unlocked_batches = ResultBatch.objects.filter(is_locked=False).count()
    locked_batches = ResultBatch.objects.filter(is_locked=True).count()

    totals = {
        "students": Student.objects.count(),
        "enrollments": Enrollment.objects.count(),
        "batches": ResultBatch.objects.count(),
        "course_results": CourseResult.objects.count(),
        "unlocked_batches": unlocked_batches,
        "locked_batches": locked_batches,
    }

    # Batches with no course results yet (helpful for "what to work on")
    todo_batches = (
        ResultBatch.objects.filter(is_locked=False)
        .annotate(cr_count=Count("course_results"))
        .filter(cr_count=0)
        .select_related("program", "session")
        .order_by("-created_at")[:8]
    )

    ctx = {
        "totals": totals,
        "todo_batches": todo_batches,
    }
    return render(request, "dashboards/data_entry.html", ctx)


def _norm(s: str) -> str:
    return "".join(ch.lower() for ch in str(s).strip() if ch.isalnum())


def _to_float_or_zero(v):
    if v is None or str(v).strip() == "":
        return 0.0
    return float(v)


@group_required("Data Entry")
@transaction.atomic
def data_entry_import_marks(request):
    """
    Upload an Excel file and import marks.

    This follows the same rules as results/management/commands/import_marks.py
    (including column matching and creating/updating CourseResult).
    """
    if request.method == "POST":
        recompute = request.POST.get("recompute") == "on"
        xlsx = request.FILES.get("file")

        if not xlsx:
            messages.error(request, "Please choose an Excel (.xlsx) file.")
            return redirect("data_entry_import_marks")

        if not xlsx.name.lower().endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect("data_entry_import_marks")

        # Save upload into MEDIA/imports
        imports_dir = os.path.join(settings.MEDIA_ROOT, "imports")
        os.makedirs(imports_dir, exist_ok=True)
        fs = FileSystemStorage(location=imports_dir)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = f"marks_{ts}_{os.path.basename(xlsx.name)}"
        filename = fs.save(safe_name, xlsx)
        file_path = fs.path(filename)

        created = 0
        updated = 0
        errors: list[str] = []
        touched_batches: dict[tuple[int, int, int, str], ResultBatch] = {}

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            header_raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            header = [_norm(h) for h in header_raw]

            def col(*names):
                for n in names:
                    n2 = _norm(n)
                    for i, h in enumerate(header):
                        if h == n2:
                            return i
                return None

            reg_i = col("registration_no", "registrationno")
            prog_i = col("program")
            sess_i = col("session")
            sem_i = col("semester")
            code_i = col("course_code", "coursecode", "code")
            title_i = col("course_title", "coursetitle", "title")

            sesm_i = col("sessional_marks", "sessional")
            mid_i = col("midterm_marks", "midterm", "mid")
            ter_i = col("terminal_marks", "terminal", "final")
            max_i = col("maxmarks", "max_marks", "max")
            exam_i = col("examtype", "result_type", "type")

            required = {
                "registration_no": reg_i,
                "program": prog_i,
                "session": sess_i,
                "semester": sem_i,
                "terminal_marks": ter_i,
                "maxmarks": max_i,
            }
            missing = [k for k, v in required.items() if v is None]
            if missing:
                messages.error(
                    request,
                    "Missing required columns: "
                    + ", ".join(missing)
                    + f". Headers found: {header_raw}",
                )
                return redirect("data_entry_import_marks")

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    registration_no = row[reg_i]
                    program_name = row[prog_i]
                    session_year = row[sess_i]
                    semester_number = row[sem_i]
                    terminal_marks = row[ter_i]
                    max_marks = row[max_i]

                    course_code = row[code_i] if code_i is not None else None
                    course_title = row[title_i] if title_i is not None else None
                    sessional_marks = row[sesm_i] if sesm_i is not None else None
                    midterm_marks = row[mid_i] if mid_i is not None else None
                    examtype = row[exam_i] if exam_i is not None else "Regular"

                    if not registration_no or not program_name or not session_year or not semester_number:
                        errors.append(f"Row {row_num}: missing program/session/semester/registration_no")
                        continue

                    if terminal_marks is None or str(terminal_marks).strip() == "":
                        errors.append(f"Row {row_num}: terminal_marks is required")
                        continue

                    if max_marks is None or str(max_marks).strip() == "":
                        errors.append(f"Row {row_num}: maxmarks is required")
                        continue

                    registration_no = str(registration_no).strip()
                    program_name = str(program_name).strip()
                    session_year = int(session_year)
                    semester_number = int(semester_number)

                    examtype = str(examtype).strip().lower()
                    if examtype in ("regular",):
                        result_type = "regular"
                    elif examtype in ("repeat", "reappear"):
                        result_type = "repeat"
                    elif examtype in ("improved", "improvement"):
                        result_type = "improved"
                    else:
                        result_type = "regular"

                    program = Program.objects.filter(name=program_name).first() or Program.objects.filter(
                        name__icontains=program_name
                    ).first()
                    if not program:
                        errors.append(f"Row {row_num}: Program not found: {program_name}")
                        continue

                    session = Session.objects.filter(start_year=session_year).first()
                    if not session:
                        errors.append(f"Row {row_num}: Session not found: {session_year}")
                        continue

                    student = Student.objects.filter(registration_no=registration_no).first()
                    if not student:
                        errors.append(f"Row {row_num}: Student not found: {registration_no}")
                        continue

                    enrollment = Enrollment.objects.filter(student=student, program=program, session=session).first()
                    if not enrollment:
                        errors.append(
                            f"Row {row_num}: Enrollment not found for reg={registration_no} program={program.name} session={session.start_year}"
                        )
                        continue

                    # Course match: by code preferred, else title
                    course = None
                    if course_code is not None and str(course_code).strip() != "":
                        course = Course.objects.filter(code=str(course_code).strip()).first()
                    if not course and course_title:
                        course = Course.objects.filter(title=str(course_title).strip()).first()
                    if not course:
                        errors.append(f"Row {row_num}: Course not found (code={course_code}, title={course_title})")
                        continue

                    key = (program.id, session.id, semester_number, result_type)
                    if key not in touched_batches:
                        batch, _ = ResultBatch.objects.get_or_create(
                            program=program,
                            session=session,
                            semester_number=semester_number,
                            result_type=result_type,
                        )
                        touched_batches[key] = batch
                    batch = touched_batches[key]

                    if batch.is_locked:
                        errors.append(f"Row {row_num}: Batch is locked: {batch}")
                        continue

                    total = (
                        _to_float_or_zero(sessional_marks)
                        + _to_float_or_zero(midterm_marks)
                        + _to_float_or_zero(terminal_marks)
                    )

                    cr = CourseResult.objects.filter(batch=batch, enrollment=enrollment, course=course).first()
                    if not cr:
                        CourseResult.objects.create(
                            batch=batch,
                            enrollment=enrollment,
                            course=course,
                            marks_obtained=total,
                            max_marks=float(max_marks),
                        )
                        created += 1
                    else:
                        cr.marks_obtained = total
                        cr.max_marks = float(max_marks)
                        cr.save(update_fields=["marks_obtained", "max_marks"])
                        updated += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: ERROR {e}")

            if recompute:
                for batch in touched_batches.values():
                    recompute_batch(batch)

        except Exception as e:
            messages.error(request, f"Import failed: {e}")
            return redirect("data_entry_import_marks")

        ctx = {
            "created": created,
            "updated": updated,
            "error_count": len(errors),
            "errors": errors[:200],
            "recompute": recompute,
            "batches": list(touched_batches.values()),
        }
        return render(request, "dashboards/data_entry_import_result.html", ctx)

    return render(request, "dashboards/data_entry_import.html")


@group_required("Document Generator")
def document_generator_dashboard(request):
    return render(request, "dashboards/document_generator.html")


@group_required("Result Checker")
def result_checker_dashboard(request):
    return render(request, "dashboards/result_checker.html")


@group_required("System Admin")
def system_admin_dashboard(request):
    return render(request, "dashboards/system_admin.html")
