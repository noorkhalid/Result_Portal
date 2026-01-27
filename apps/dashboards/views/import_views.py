from __future__ import annotations

import os
from datetime import datetime

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render

from academics.models import Course, Program, Session, Department, ProgramCourse
from students.models import Student, Enrollment
from dashboards.decorators import group_required


def _norm(v):
    return str(v or "").strip().lower().replace(" ", "_")


def _first_sheet(wb: openpyxl.Workbook):
    return wb.active


def _col_index(header, *names):
    for name in names:
        key = _norm(name)
        if key in header:
            return header.index(key)
    return None


def _bool(v, default=True):
    if v is None or str(v).strip() == "":
        return default
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "active")


def _decimal(v):
    if v is None or str(v).strip() == "":
        return None
    return float(v)


# ======================================================
# EXCEL TEMPLATES (DOWNLOAD)
# ======================================================

@group_required("System Admin")
def template_courses(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses"
    ws.append(["code", "title", "credit_hours"])
    ws.append(["CS101", "Introduction to Computing", 3])

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="courses_template.xlsx"'
    wb.save(resp)
    return resp


@group_required("System Admin")
def template_students(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["registration_no", "name", "father_name", "is_active"])
    ws.append(["2021-ABC-001", "Ali Khan", "Ahmed Khan", True])

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="students_template.xlsx"'
    wb.save(resp)
    return resp


@group_required("System Admin")
def template_enrollments(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrollments"
    ws.append(["registration_no", "program", "session", "roll_no", "is_active"])
    ws.append(["2021-ABC-001", "BS Computer Science", 2021, "BSCS-001", True])

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="enrollments_template.xlsx"'
    wb.save(resp)
    return resp


@group_required("System Admin")
def template_program_courses(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ProgramCourses"
    ws.append(["department", "program", "semester_number", "course_code"])
    ws.append(["Falcon Educational Complex, Tank", "BS Computer Science", 1, "CS101"])

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="program_courses_template.xlsx"'
    wb.save(resp)
    return resp



# ======================================================
# IMPORT COURSES
# ======================================================

@group_required("System Admin")
@transaction.atomic

@group_required("System Admin")
@transaction.atomic
@group_required("System Admin")
def import_courses(request):
    """
    Import Courses from Excel (.xlsx).

    Rules:
    - Courses are GLOBAL (not department-bound).
    - Course.code is UNIQUE (case-insensitive match is used during import).
    - Course.title is NOT unique.
    - All-or-nothing: if ANY row has an error, NOTHING is imported.
    """
    if request.method != "POST":
        return render(request, "dashboards/imports/courses_import.html")

    xlsx = request.FILES.get("file")
    if not xlsx:
        messages.error(request, "Please choose an Excel (.xlsx) file.")
        return redirect("admin_import_courses")

    if not xlsx.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx files are supported.")
        return redirect("admin_import_courses")

    imports_dir = os.path.join(settings.MEDIA_ROOT, "imports")
    os.makedirs(imports_dir, exist_ok=True)
    fs = FileSystemStorage(location=imports_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = fs.save(f"courses_{ts}_{xlsx.name}", xlsx)
    file_path = fs.path(filename)

    required_cols = {"code", "title", "credit_hours"}

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows or not rows[0]:
            messages.error(request, "Excel file is empty.")
            return redirect("admin_import_courses")

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        missing = required_cols - set(header)
        if missing:
            messages.error(request, f"Missing required columns: {', '.join(sorted(missing))}.")
            return redirect("admin_import_courses")

        idx = {name: header.index(name) for name in required_cols}

        # Pre-validate all rows (no DB writes yet)
        cleaned = []
        seen_codes = set()
        validation_errors = []

        for i, row in enumerate(rows[1:], start=2):  # Excel row numbers start at 1; header is row 1
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue  # skip completely empty rows

            code = (row[idx["code"]] or "")
            title = (row[idx["title"]] or "")
            ch = row[idx["credit_hours"]]

            code = str(code).strip()
            title = str(title).strip()

            if not code:
                validation_errors.append(f"Row {i}: code is required.")
                continue
            if not title:
                validation_errors.append(f"Row {i}: title is required.")
                continue

            code_key = code.lower()
            if code_key in seen_codes:
                validation_errors.append(f"Row {i}: duplicate code '{code}' in the same file.")
                continue
            seen_codes.add(code_key)

            try:
                # allow int/float/str numeric
                credit_hours = float(ch) if isinstance(ch, (int, float)) else float(str(ch).strip())
            except Exception:
                validation_errors.append(f"Row {i}: invalid credit_hours '{ch}'.")
                continue

            cleaned.append((code, title, credit_hours))

        if validation_errors:
            messages.error(request, "Import failed. Fix these errors and try again:\n" + "\n".join(validation_errors[:50]))
            return redirect("admin_import_courses")

        # All-or-nothing DB write
        created = 0
        updated = 0
        with transaction.atomic():
            for code, title, credit_hours in cleaned:
                obj = Course.objects.filter(code__iexact=code).first()
                if obj:
                    obj.code = code  # normalize casing
                    obj.title = title
                    obj.credit_hours = credit_hours
                    obj.save()
                    updated += 1
                else:
                    Course.objects.create(code=code, title=title, credit_hours=credit_hours)
                    created += 1

        messages.success(request, f"Courses imported successfully: {created} created, {updated} updated.")
        return redirect("admin_course_list")

    except Exception as e:
        messages.error(request, f"Failed to import courses: {e}")
        return redirect("admin_import_courses")




# ======================================================
# IMPORT STUDENTS
# ======================================================

@group_required("System Admin")
@transaction.atomic
def import_students(request):
    if request.method != "POST":
        return render(request, "dashboards/imports/students_import.html")

    xlsx = request.FILES.get("file")
    if not xlsx:
        messages.error(request, "Please choose an Excel (.xlsx) file.")
        return redirect("admin_import_students")

    if not xlsx.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx files are supported.")
        return redirect("admin_import_students")

    imports_dir = os.path.join(settings.MEDIA_ROOT, "imports")
    os.makedirs(imports_dir, exist_ok=True)
    fs = FileSystemStorage(location=imports_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = fs.save(f"students_{ts}_{xlsx.name}", xlsx)
    file_path = fs.path(filename)

    created = 0
    updated = 0
    errors: list[str] = []

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = _first_sheet(wb)

        header_raw = [c.value for c in ws[1]]
        header = [_norm(h) for h in header_raw]

        reg_i = _col_index(header, "registration_no", "reg_no", "registration")
        name_i = _col_index(header, "name", "student_name")
        father_i = _col_index(header, "father_name", "father")
        active_i = _col_index(header, "is_active", "active")

        required = {"registration_no": reg_i, "name": name_i, "father_name": father_i}
        missing = [k for k, v in required.items() if v is None]
        if missing:
            messages.error(request, f"Missing columns: {', '.join(missing)}")
            return redirect("admin_import_students")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            registration_no = str(row[reg_i] or "").strip()
            name = str(row[name_i] or "").strip()
            father_name = str(row[father_i] or "").strip()
            is_active = _bool(row[active_i], default=True) if active_i is not None else True

            if not registration_no or not name or not father_name:
                errors.append(f"Row {row_num}: missing required values")
                continue

            obj = Student.objects.filter(registration_no__iexact=registration_no).first()
            if obj:
                obj.name = name
                obj.father_name = father_name
                obj.is_active = is_active
                obj.save()
                updated += 1
            else:
                try:
                    Student.objects.create(
                        registration_no=registration_no,
                        name=name,
                        father_name=father_name,
                        is_active=is_active,
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

    except Exception as e:
        messages.error(request, f"Failed to import students: {e}")
        return redirect("admin_import_students")

    if created or updated:
        messages.success(request, f"Students imported: {created} created, {updated} updated.")
    if errors:
        messages.warning(request, "Some rows were skipped:\n" + "\n".join(errors[:25]))

    return redirect("admin_student_list")


# ======================================================
# IMPORT ENROLLMENTS
# ======================================================

@group_required("System Admin")
@transaction.atomic
def import_enrollments(request):
    if request.method != "POST":
        return render(request, "dashboards/imports/enrollments_import.html")

    xlsx = request.FILES.get("file")
    if not xlsx:
        messages.error(request, "Please choose an Excel (.xlsx) file.")
        return redirect("admin_import_enrollments")

    if not xlsx.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx files are supported.")
        return redirect("admin_import_enrollments")

    imports_dir = os.path.join(settings.MEDIA_ROOT, "imports")
    os.makedirs(imports_dir, exist_ok=True)
    fs = FileSystemStorage(location=imports_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = fs.save(f"enrollments_{ts}_{xlsx.name}", xlsx)
    file_path = fs.path(filename)

    created = 0
    updated = 0
    errors: list[str] = []

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = _first_sheet(wb)

        header_raw = [c.value for c in ws[1]]
        header = [_norm(h) for h in header_raw]

        reg_i = _col_index(header, "registration_no", "reg_no", "registration")
        prog_i = _col_index(header, "program", "program_name")
        sess_i = _col_index(header, "session", "session_year", "start_year")
        roll_i = _col_index(header, "roll_no", "rollno", "roll_number")
        active_i = _col_index(header, "is_active", "active")

        required = {"registration_no": reg_i, "program": prog_i, "session": sess_i, "roll_no": roll_i}
        missing = [k for k, v in required.items() if v is None]
        if missing:
            messages.error(request, f"Missing columns: {', '.join(missing)}")
            return redirect("admin_import_enrollments")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            registration_no = str(row[reg_i] or "").strip()
            program_name = str(row[prog_i] or "").strip()
            session_year = row[sess_i]
            roll_no = str(row[roll_i] or "").strip()
            is_active = _bool(row[active_i], default=True) if active_i is not None else True

            if not registration_no or not program_name or session_year is None or not roll_no:
                errors.append(f"Row {row_num}: missing required values")
                continue

            try:
                session_year_int = int(session_year)
            except Exception:
                errors.append(f"Row {row_num}: invalid session year '{session_year}'")
                continue

            student = Student.objects.filter(registration_no__iexact=registration_no).first()
            if not student:
                errors.append(f"Row {row_num}: student not found ({registration_no})")
                continue

            program = Program.objects.filter(name__icontains=program_name).first()
            if not program:
                errors.append(f"Row {row_num}: program not found ({program_name})")
                continue

            session = Session.objects.filter(start_year=session_year_int).first()
            if not session:
                errors.append(f"Row {row_num}: session not found ({session_year_int})")
                continue

            obj = Enrollment.objects.filter(student=student, program=program, session=session).first()
            if not obj:
                obj = Enrollment.objects.filter(program=program, session=session, roll_no=roll_no).first()

            if obj:
                obj.student = student
                obj.program = program
                obj.session = session
                obj.roll_no = roll_no
                obj.is_active = is_active
                obj.save()
                updated += 1
            else:
                try:
                    Enrollment.objects.create(
                        student=student,
                        program=program,
                        session=session,
                        roll_no=roll_no,
                        is_active=is_active,
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

    except Exception as e:
        messages.error(request, f"Failed to import enrollments: {e}")
        return redirect("admin_import_enrollments")

    if created or updated:
        messages.success(request, f"Enrollments imported: {created} created, {updated} updated.")
    if errors:
        messages.warning(request, "Some rows were skipped:\n" + "\n".join(errors[:25]))

    return redirect("admin_enrollment_list")


# ======================================================
# IMPORT PROGRAM COURSES (MAPPING)
# ======================================================

@group_required("System Admin")
@transaction.atomic
def import_program_courses(request):
    """
    Import ProgramCourse mappings from Excel (.xlsx).

    Columns required:
    - department: Department name (must exist)
    - program: Program name (must exist; matched within department)
    - semester_number: integer
    - course_code: existing Course.code

    All-or-nothing: if ANY row has an error, NOTHING is imported.
    """
    if request.method != "POST":
        return render(request, "dashboards/imports/program_courses_import.html")

    xlsx = request.FILES.get("file")
    if not xlsx:
        messages.error(request, "Please choose an Excel (.xlsx) file.")
        return redirect("admin_import_program_courses")

    if not xlsx.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx files are supported.")
        return redirect("admin_import_program_courses")

    imports_dir = os.path.join(settings.MEDIA_ROOT, "imports")
    os.makedirs(imports_dir, exist_ok=True)
    fs = FileSystemStorage(location=imports_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = fs.save(f"program_courses_{ts}_{xlsx.name}", xlsx)
    file_path = fs.path(filename)

    required_cols = {"department", "program", "semester_number", "course_code"}

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows or not rows[0]:
            messages.error(request, "Excel file is empty.")
            return redirect("admin_import_program_courses")

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        missing = required_cols - set(header)
        if missing:
            messages.error(request, f"Missing required columns: {', '.join(sorted(missing))}.")
            return redirect("admin_import_program_courses")

        idx = {name: header.index(name) for name in required_cols}

        cleaned = []
        validation_errors = []

        for i, row in enumerate(rows[1:], start=2):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue

            dept_name = str(row[idx["department"]] or "").strip()
            program_name = str(row[idx["program"]] or "").strip()
            sem_raw = row[idx["semester_number"]]
            course_code = str(row[idx["course_code"]] or "").strip()

            if not dept_name:
                validation_errors.append(f"Row {i}: department is required.")
                continue
            if not program_name:
                validation_errors.append(f"Row {i}: program is required.")
                continue
            if sem_raw is None or str(sem_raw).strip() == "":
                validation_errors.append(f"Row {i}: semester_number is required.")
                continue
            if not course_code:
                validation_errors.append(f"Row {i}: course_code is required.")
                continue

            department = Department.objects.filter(name__iexact=dept_name).first()
            if not department:
                validation_errors.append(f"Row {i}: department not found ('{dept_name}').")
                continue

            program = Program.objects.filter(department=department, name__iexact=program_name).first()
            if not program:
                program = Program.objects.filter(department=department, name__icontains=program_name).first()
            if not program:
                validation_errors.append(f"Row {i}: program not found in department ('{program_name}').")
                continue

            try:
                semester_number = int(sem_raw)
            except Exception:
                validation_errors.append(f"Row {i}: invalid semester_number '{sem_raw}'.")
                continue

            course = Course.objects.filter(code__iexact=course_code).first()
            if not course:
                validation_errors.append(f"Row {i}: course not found by code ('{course_code}').")
                continue

            cleaned.append((department, program, semester_number, course))

        if validation_errors:
            messages.error(
                request,
                "Import failed. Fix these errors and try again:\n" + "\n".join(validation_errors[:50]),
            )
            return redirect("admin_import_program_courses")

        created = 0
        updated = 0

        with transaction.atomic():
            for department, program, semester_number, course in cleaned:
                obj = ProgramCourse.objects.filter(
                    program=program, semester_number=semester_number, course=course
                ).first()
                if obj:
                    obj.department = department
                    obj.save()
                    updated += 1
                else:
                    ProgramCourse.objects.create(
                        department=department,
                        program=program,
                        semester_number=semester_number,
                        course=course,
                    )
                    created += 1

        messages.success(
            request,
            f"Program Courses imported successfully: {created} created, {updated} updated.",
        )
        return redirect("admin_program_course_list")

    except Exception as e:
        messages.error(request, f"Failed to import program courses: {e}")
        return redirect("admin_import_program_courses")
