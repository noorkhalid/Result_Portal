from __future__ import annotations

import os
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.db.models import Count
from django.shortcuts import redirect, render
from django.http import HttpResponse

from academics.models import Course, Program, Session
from results.models import CourseResult, ExamType, ResultBatch
from results.services import recompute_batch
from students.models import Enrollment, Student

from dashboards.decorators import group_required


# ======================================================
# PUBLIC / ENTRY
# ======================================================

def home(request):
    """Public entry point.
    If logged in → route to dashboard
    Else → login
    """
    if request.user.is_authenticated:
        return redirect("dashboard")
    return redirect("login")


# ======================================================
# DASHBOARD ROUTER
# ======================================================

@login_required
def dashboard(request):
    """Route the logged-in user to the correct dashboard.
    System Admin ALWAYS has priority.
    """
    user = request.user

    # Highest priority: System Admin
    if user.is_superuser or user.groups.filter(name="System Admin").exists():
        return redirect("dash_system_admin")

    if user.groups.filter(name="Controller").exists():
        return redirect("dash_controller")

    if user.groups.filter(name="Data Entry").exists():
        return redirect("dash_data_entry")

    if user.groups.filter(name="Document Generator").exists():
        return redirect("dash_document_generator")

    if user.groups.filter(name="Result Checker").exists():
        return redirect("dash_result_checker")

    return render(request, "dashboards/no_group.html")


# ======================================================
# ROLE DASHBOARDS
# ======================================================

@group_required("Controller")
def controller_dashboard(request):
    return render(request, "dashboards/controller.html")


@group_required("Data Entry")
def data_entry_dashboard(request):
    """Data Entry dashboard with quick stats and pending batches."""
    unlocked_batches = ResultBatch.objects.filter(is_locked=False).count()
    locked_batches = ResultBatch.objects.filter(is_locked=True).count()

    totals = {
        "students": Student.objects.count(),
        "enrollments": Enrollment.objects.count(),
        "batches": ResultBatch.objects.count(),
        "course_results": CourseResult.objects.count(),
        "unlocked_batches": unlocked_batches,
        "locked_batches": locked_batches,
    }

    todo_batches = (
        ResultBatch.objects.filter(is_locked=False)
        .annotate(cr_count=Count("course_results"))
        .filter(cr_count=0)
        .select_related("program", "session")
        .order_by("-created_at")[:8]
    )

    return render(
        request,
        "dashboards/data_entry.html",
        {"totals": totals, "todo_batches": todo_batches},
    )


@group_required("Document Generator")
def document_generator_dashboard(request):
    return render(request, "dashboards/document_generator.html")


@group_required("Result Checker")
def result_checker_dashboard(request):
    return render(request, "dashboards/result_checker.html")


@group_required("System Admin")
def system_admin_dashboard(request):
    """System Admin landing.

    The old "System Admin Dashboard" page (cards/stats) has been removed.
    Keep this URL as a stable entry point and redirect to the first working page.
    """
    return redirect("admin_program_list")


# ======================================================
# DATA ENTRY — EXCEL IMPORT
# ======================================================

def _norm(s: str) -> str:
    """Normalize header strings."""
    return "".join(ch.lower() for ch in str(s).strip() if ch.isalnum())


def _to_float_or_zero(v):
    if v is None or str(v).strip() == "":
        return 0.0
    return float(v)


def _to_int_or_none(v):
    """Safe int parsing for Excel cells."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    try:
        # Excel sometimes provides numeric strings or floats.
        return int(float(v))
    except Exception:
        return None


def _to_decimal_2_or_none(v):
    """Parse to Decimal with exactly 2 decimal places, or None if blank/invalid."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    try:
        d = Decimal(str(v))
    except (InvalidOperation, ValueError, TypeError):
        return None
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _to_decimal_2_or_zero(v):
    d = _to_decimal_2_or_none(v)
    return d if d is not None else Decimal("0.00")


@group_required("Data Entry")
@transaction.atomic
def data_entry_import_marks(request):
    """Upload Excel (.xlsx) file and import marks."""
    if request.method != "POST":
        return render(request, "dashboards/data_entry_import.html")

    recompute = request.POST.get("recompute") == "on"
    xlsx = request.FILES.get("file")

    if not xlsx:
        messages.error(request, "Please choose an Excel (.xlsx) file.")
        return redirect("data_entry_import_marks")

    if not xlsx.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx files are supported.")
        return redirect("data_entry_import_marks")

    imports_dir = os.path.join(settings.MEDIA_ROOT, "imports")
    os.makedirs(imports_dir, exist_ok=True)

    fs = FileSystemStorage(location=imports_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = fs.save(f"marks_{ts}_{xlsx.name}", xlsx)
    file_path = fs.path(filename)

    created = 0
    updated = 0
    errors: list[str] = []
    touched_batches: dict[tuple[int, int, int, str], ResultBatch] = {}

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        header_raw = [c.value for c in ws[1]]
        header = [_norm(h) for h in header_raw]

        def col(*names):
            for name in names:
                key = _norm(name)
                if key in header:
                    return header.index(key)
            return None

        reg_i = col("registration_no")
        prog_i = col("program")
        sess_i = col("session")
        sem_i = col("semester")
        code_i = col("course_code", "code")
        title_i = col("course_title", "title")
        # Accept common template header variants (older templates used different names)
        ses_i = col("sessional_marks", "sessional", "sessiononal", "sessional_")
        mid_i = col("midterm_marks", "midterm", "midterm_")
        ter_i = col("terminal_marks", "terminal", "terminal_")
        max_i = col("maxmarks")
        exam_i = col("examtype")

        required = {
            "registration_no": reg_i,
            "program": prog_i,
            "session": sess_i,
            "semester": sem_i,
            "terminal_marks": ter_i,
            "maxmarks": max_i,
        }

        missing = [k for k, v in required.items() if v is None]
        if missing:
            messages.error(request, f"Missing columns: {', '.join(missing)}")
            return redirect("data_entry_import_marks")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                registration_no = str(row[reg_i]).strip()
                program_name = str(row[prog_i]).strip()
                session_year = _to_int_or_none(row[sess_i])
                semester_number = _to_int_or_none(row[sem_i])

                terminal = _to_decimal_2_or_none(row[ter_i])
                max_marks = _to_decimal_2_or_none(row[max_i])

                if session_year is None or semester_number is None:
                    errors.append(f"Row {row_num}: invalid session/semester")
                    continue

                if terminal is None or max_marks is None:
                    errors.append(f"Row {row_num}: missing marks")
                    continue

                program = Program.objects.filter(name__icontains=program_name).first()
                session = Session.objects.filter(start_year=session_year).first()
                student = Student.objects.filter(registration_no=registration_no).first()

                if not all([program, session, student]):
                    errors.append(f"Row {row_num}: invalid program/session/student")
                    continue

                enrollment = Enrollment.objects.filter(
                    student=student, program=program, session=session
                ).first()

                if not enrollment:
                    errors.append(f"Row {row_num}: enrollment not found")
                    continue

                course = None
                if code_i is not None and row[code_i]:
                    course = Course.objects.filter(code=str(row[code_i]).strip()).first()
                if not course and title_i is not None and row[title_i]:
                    course = Course.objects.filter(title=str(row[title_i]).strip()).first()

                if not course:
                    errors.append(f"Row {row_num}: course not found")
                    continue

                examtype_code = str(row[exam_i]).strip().lower() if exam_i is not None and row[exam_i] else "regular"
                # keep backward compatibility with existing templates/imports
                if not examtype_code:
                    examtype_code = "regular"

                exam_type, _ = ExamType.objects.get_or_create(
                    code=examtype_code,
                    defaults={"name": examtype_code.replace("_", " ").replace("-", " ").title()},
                )

                # Batch uniqueness is per-department.
                dept_id = enrollment.department_id
                key = (dept_id, program.id, session.id, semester_number, exam_type.id)
                batch = touched_batches.get(key)

                if not batch:
                    batch, _ = ResultBatch.objects.get_or_create(
                        department_id=dept_id,
                        program=program,
                        session=session,
                        semester_number=semester_number,
                        exam_type=exam_type,
                    )
                    touched_batches[key] = batch

                if batch.is_locked:
                    errors.append(f"Row {row_num}: batch locked")
                    continue

                ses_val = _to_decimal_2_or_none(row[ses_i] if ses_i is not None else None)
                mid_val = _to_decimal_2_or_none(row[mid_i] if mid_i is not None else None)
                ter_val = terminal

                total = (
                    (ses_val or Decimal("0.00"))
                    + (mid_val or Decimal("0.00"))
                    + (ter_val or Decimal("0.00"))
                ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

                cr, created_flag = CourseResult.objects.get_or_create(
                    batch=batch,
                    enrollment=enrollment,
                    course=course,
                    defaults={
                        "sessional_marks": ses_val,
                        "midterm_marks": mid_val,
                        "terminal_marks": ter_val,
                        "marks_obtained": total,
                        "max_marks": max_marks,
                    },
                )

                if not created_flag:
                    cr.sessional_marks = ses_val
                    cr.midterm_marks = mid_val
                    cr.terminal_marks = ter_val
                    cr.marks_obtained = total
                    cr.max_marks = max_marks
                    cr.save(update_fields=[
                        "sessional_marks",
                        "midterm_marks",
                        "terminal_marks",
                        "marks_obtained",
                        "max_marks",
                    ])
                    updated += 1
                else:
                    created += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {e}")

        if recompute:
            for batch in touched_batches.values():
                recompute_batch(batch)

    except Exception as e:
        messages.error(request, f"Import failed: {e}")
        return redirect("data_entry_import_marks")

    return render(
        request,
        "dashboards/data_entry_import_result.html",
        {
            "created": created,
            "updated": updated,
            "error_count": len(errors),
            "errors": errors[:200],
            "batches": list(touched_batches.values()),
            "recompute": recompute,
        },
    )


@group_required("Data Entry")
def data_entry_marks_template(request):
    """Download an Excel template for marks import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks"
    ws.append([
        "registration_no",
        "program",
        "session",
        "semester",
        "course_code",
        "course_title",
        "sessional_marks",
        "midterm_marks",
        "terminal_marks",
        "maxmarks",
        "examtype",
    ])
    ws.append(["2021-ABC-001", "BS Computer Science", 2021, 1, "CS101", "Introduction to Computing", 10, 20, 50, 100, "Regular"])

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="marks_template.xlsx"'
    wb.save(resp)
    return resp
