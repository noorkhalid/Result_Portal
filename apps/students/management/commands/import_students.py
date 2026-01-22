from django.core.management.base import BaseCommand
from students.models import Student, Enrollment
from academics.models import Program, Session
import openpyxl


def _norm(s: str) -> str:
    return "".join(ch.lower() for ch in str(s).strip() if ch.isalnum())


class Command(BaseCommand):
    help = "Import students + enrollments from Excel."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")
        parser.add_argument("--program", required=True, help="Program name (match or partial match)")
        parser.add_argument("--session", required=True, type=int, help="Session start year e.g. 2022")

    def handle(self, *args, **options):
        file_path = options["file"]
        program_text = str(options["program"]).strip()
        session_year = options["session"]

        # Program fuzzy match
        program = Program.objects.filter(name=program_text).first()
        if not program:
            program = Program.objects.filter(name__icontains=program_text).first()
        if not program:
            raise SystemExit(f"Program not found: {program_text}")

        session = Session.objects.filter(start_year=session_year).first()
        if not session:
            raise SystemExit(f"Session not found: {session_year}")

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        header_raw = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        header = [_norm(h) for h in header_raw]

        def find_col(candidates):
            for cand in candidates:
                cand_n = _norm(cand)
                for i, h in enumerate(header):
                    if h == cand_n:
                        return i
            return None

        roll_idx = find_col(["roll_no", "rollno", "roll", "rollnumber"])
        reg_idx = find_col(["registration_no", "registrationno", "regno", "reg_no", "registration"])
        name_idx = find_col(["name", "studentname", "student_name"])
        father_idx = find_col(["father_name", "fathername", "fname", "father"])

        missing = []
        if roll_idx is None: missing.append("roll_no")
        if reg_idx is None: missing.append("registration_no")
        if name_idx is None: missing.append("name")
        if father_idx is None: missing.append("father_name")

        if missing:
            raise SystemExit(
                f"Missing columns in Excel: {missing}\n"
                f"Your headers are: {header_raw}"
            )

        created_students = 0
        updated_students = 0
        created_enroll = 0
        updated_enroll = 0
        errors = 0

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            roll_no = row[roll_idx] if roll_idx < len(row) else None
            reg_no = row[reg_idx] if reg_idx < len(row) else None
            name = row[name_idx] if name_idx < len(row) else None
            father = row[father_idx] if father_idx < len(row) else None

            if not roll_no or not reg_no or not name or not father:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Row {row_num}: Missing required fields (skipped)"))
                continue

            roll_no = str(roll_no).strip()
            reg_no = str(reg_no).strip()
            name = str(name).strip()
            father = str(father).strip()

            student = Student.objects.filter(registration_no=reg_no).first()
            if not student:
                student = Student.objects.create(
                    registration_no=reg_no,
                    name=name,
                    father_name=father,
                )
                created_students += 1
            else:
                changed = False
                if student.name != name:
                    student.name = name
                    changed = True
                if student.father_name != father:
                    student.father_name = father
                    changed = True
                if changed:
                    student.save(update_fields=["name", "father_name"])
                    updated_students += 1

            enroll = Enrollment.objects.filter(program=program, session=session, roll_no=roll_no).first()
            if not enroll:
                Enrollment.objects.create(
                    student=student,
                    program=program,
                    session=session,
                    roll_no=roll_no,
                )
                created_enroll += 1
            else:
                if enroll.student_id != student.id:
                    enroll.student = student
                    enroll.save(update_fields=["student"])
                    updated_enroll += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nDone. Students: created={created_students}, updated={updated_students} | "
            f"Enrollments: created={created_enroll}, updated={updated_enroll} | errors={errors}"
        ))
